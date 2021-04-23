import os
import xlwt
import openpyxl

""" 读取原始数据： eyeTrack_dict以字典的格式储存， eyeTrack_dict[顺序编号] = list([可视化区域, 持续时间]) """
def loadEyeTrackerData(file_path):
    print("begin reading eye tracker data of ", file_path)
    eyeTrack_dict = {}
    with open(file_path, "r", encoding="utf8", errors="ignore") as file:
        for i, line in enumerate(file.readlines()):
            split = line.strip("\n").split(",")
            if len(split) != 3:
                print("This %d line - %s - is in the wrong format"%(i, line.strip("\n")))
            else:
                category, duration = str(split[0]), int(split[-1])
                eyeTrack_dict[i] = list([category, duration])
    return eyeTrack_dict


def outputProbData(dir_path, file_name, data_dict):
    dir_path = str(dir_path)
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    with open(dir_path + file_name, "w", encoding="utf8", errors="ignore") as file:
        for key in data_dict.keys():
            file.write(data_dict[key][0] + "," + str(data_dict[key][-1]))
            file.write("\n")



""" 两种数据构造方式：
 1. 将“注视”信息按照 时间/总长 的方式转变成概率数据；
 2. 将“注视转移”信息按照 注视A/注视B 的方式转变成概率数据。
 """

"""  1. 将“注视”信息按照 时间/总长 的方式转变成概率数据 """
def crateGazeProbData(raw_data):
    get_duration = lambda item: item[-1]
    total_duration = sum([get_duration(item) for item in list(raw_data.values())])
    Gaze_Prob_Data = {id: [raw_data[id][0], raw_data[id][-1]/total_duration] for id in raw_data.keys()}
    # print(Gaze_Prob_Data)
    return Gaze_Prob_Data

"""  2. 将“注视转移”信息按照 注视A/注视B 的方式转变成概率数据。 """
def crateGazeTrackProbData(raw_data):
    get_duration = lambda item: item[-1]
    total_duration = sum([get_duration(item) for item in list(raw_data.values())])
    key_list = sorted(list(raw_data.keys()))
    #print("raw_data")
    #print(raw_data)
    #print("key_list")
    #print(key_list)
    GazeTrack_Prob_Data = {key:["%s->%s"%(raw_data[key_list[i-1]][0], raw_data[key][0]), (raw_data[key_list[i-1]][1] + raw_data[key][1]) / (total_duration*2-raw_data[key_list[0]][1]-raw_data[key_list[-1]][1])] for i, key in enumerate(key_list) if i > 0}
    #print("GazeTrack_Prob_Data")
    #print(GazeTrack_Prob_Data)
    return GazeTrack_Prob_Data


"""  3.  长度可变的  将“注视转移”信息按照 注视A/注视B 的方式转变成概率数据。 """
def crateGazeTrackProbData_new(raw_data, my_length):
    get_duration = lambda item: item[-1]
    total_duration = sum([get_duration(item) for item in list(raw_data.values())])
    key_list = sorted(list(raw_data.keys()))

    if len(key_list) < my_length:
        return {}

    muti_total_time = total_duration * my_length
    i = 1
    while i < my_length:
        temp = (my_length - i) * (raw_data[key_list[i - 1]][1] + raw_data[key_list[-i]][1])
        muti_total_time -= temp
        i += 1

    #print("muti_total_time")
    #print(muti_total_time)
    GazeTrack_Prob_Data = {}

    i = 0
    while i < len(key_list) - my_length + 1:# 5 - 3 + 1 = 3
        j = 0
        name = ""
        while j < my_length - 1:
            name += raw_data[i+j][0]+"->"
            j += 1
        name += raw_data[key_list[i + my_length - 1]][0]
        #print(name)
        p = 0
        for k in range(my_length):
            p += raw_data[key_list[i + k]][1]
        p = p / muti_total_time
        GazeTrack_Prob_Data[i] = [name, p]
        i += 1

    #print("GazeTrack_Prob_Data")
    #print(GazeTrack_Prob_Data)
    return GazeTrack_Prob_Data


def crateXlsFile(dir_path, sheet_name_list):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    workbook = xlwt.Workbook()
    sheet_dict = {}
    for sheet_name in sheet_name_list:
        worksheet = workbook.add_sheet(sheet_name)
        sheet_dict[sheet_name] = worksheet
    return workbook, sheet_dict


def outputPatternToXlwt(worksheet , Pattern_Summary_Dict):
    from miningFuntion import Classical_Sequential_Pattern
    worksheet.write(0, 0, "user_name")
    worksheet.write(0, 1, "pattern_name")
    worksheet.write(0, 2, "length")
    worksheet.write(0, 3, "support")
    """ pattern_name, length, support """
    i = 0
    for user_id in Pattern_Summary_Dict.keys():
        for j, key in enumerate(Pattern_Summary_Dict[user_id].keys()):
            worksheet.write(i+1, 0, user_id)
            worksheet.write(i+1, 1, key)
            worksheet.write(i+1, 2, Pattern_Summary_Dict[user_id][key].length)
            worksheet.write(i+1, 3, Pattern_Summary_Dict[user_id][key].support)
            i = i + 1


def crateXlsxFile_mutiSheet(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    workbook = openpyxl.Workbook()

    ws = workbook["Sheet"]
    workbook.remove(ws)
    return workbook

def outputRankToXlwt(book, incorrect_rank_correct_rank, sheet_name, threshold):
    worksheet = book.create_sheet(sheet_name)
    worksheet.cell(1, 1, "pattern_name")
    if 'in' in sheet_name:
        worksheet.cell(1, 2, "incorrect_rank")
        worksheet.cell(1, 3, "correct_rank")
    else:
        worksheet.cell(1, 2, "correct_rank")
        worksheet.cell(1, 3, "incorrect_rank")
    worksheet.cell(1, 4, "difference_value")
    worksheet.cell(1, 5, ">threshold="+str(threshold))
    index = 2
    for tuple in incorrect_rank_correct_rank:
        if len(tuple) == 4:
            worksheet.cell(index, 1, tuple[0])
            worksheet.cell(index, 2, tuple[1])
            worksheet.cell(index, 3, tuple[2])
            worksheet.cell(index, 4, tuple[3])
        elif len(tuple) == 5:
            worksheet.cell(index, 1, tuple[0])
            worksheet.cell(index, 2, tuple[1])
            worksheet.cell(index, 3, tuple[2])
            worksheet.cell(index, 4, tuple[3])
            worksheet.cell(index, 5, tuple[4])
        index += 1
    return
def outputAllinTopNToXlwt(book, all_in_topN):
    worksheet = book.create_sheet('all_in_topN')
    worksheet.cell(1, 1, "pattern_name")
    worksheet.cell(1, 2, "incorrect_rank")
    worksheet.cell(1, 3, "correct_rank")
    worksheet.cell(1, 4, "difference_value")
    index = 2
    for tuple in all_in_topN:
        worksheet.cell(index, 1, tuple[0])
        worksheet.cell(index, 2, tuple[1])
        worksheet.cell(index, 3, tuple[2])
        worksheet.cell(index, 4, tuple[3])
        index += 1
    return

def outputPatternToXlwt_mutiSheet(book, Pattern_Summary_Dict):
    from miningFuntion import Classical_Sequential_Pattern
    worksheet = book.create_sheet("total")  # , cell_overwrite_ok = False
    worksheet.cell(1, 1, "pattern_name")
    worksheet.cell(1, 2, "length")
    worksheet.cell(1, 3, "support")
    i = 1
    for j, key in enumerate(Pattern_Summary_Dict["total"].keys()):
        worksheet.cell(i + 1, 1, key)
        worksheet.cell(i + 1, 2, Pattern_Summary_Dict["total"][key].length)
        worksheet.cell(i + 1, 3, Pattern_Summary_Dict["total"][key].support)
        k = 4
        if len(Pattern_Summary_Dict["total"][key].username_support_list) != 0:
            for u_s in Pattern_Summary_Dict["total"][key].username_support_list:
                worksheet.cell(i + 1, k, u_s[0])
                worksheet.cell(i + 1, k + 1, u_s[1])
                k += 2
        i = i + 1

    """ pattern_name, length, support """
    for user_id in Pattern_Summary_Dict.keys():
        if user_id == "total":
            continue
        worksheet = book.create_sheet(user_id)#, cell_overwrite_ok = False
        worksheet.cell(1, 1, "pattern_name")
        worksheet.cell(1, 2, "length")
        worksheet.cell(1, 3, "support")
        i = 1
        for j, key in enumerate(Pattern_Summary_Dict[user_id].keys()):
            worksheet.cell(i+1, 1, key)
            worksheet.cell(i+1, 2, Pattern_Summary_Dict[user_id][key].length)
            worksheet.cell(i+1, 3, Pattern_Summary_Dict[user_id][key].support)
            i = i + 1
